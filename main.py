import pandas as pd

pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)

full_data = pd.read_excel('resourse.xlsx', sheet_name='eddt_5')
full_data = full_data.dropna(axis=1, how='all')  # 删除空白列
full_data = full_data.dropna(thresh=6)
# print(full_data)
full_data.to_excel('turn.xlsx', index=False)
import openpyxl

wb = openpyxl.load_workbook(r'turn.xlsx')
# print(wb)
sh = wb.active
workbook2 = openpyxl.Workbook()  # 创建一个Workbook对象，相当于创建了一个Excel文件# wb=openpyxl.Workbook(encoding='UTF-8')
worksheet2 = workbook2.active  # 获取当前活跃的worksheet,默认就是第一个worksheet
for i in range(2, len(full_data) + 2):
    for j in range(1, full_data.columns.size + 1):
        # print(full_data.columns.size)
        c = sh.cell(i, j).value
        # print(c)
        cc = sh.cell(row=i, column=j)
        # print(cc)
        # print(full_data.columns)
        # for cell in list(full_data.columns)[j - 1]:  # 获取第j列的数据
        if type(c) == str or type(c) == int:
            if j == 1:
                d = len(c)
                if d >= 4:
                    worksheet2.cell(i, 1, c)
                    ee = 'all'
                    worksheet2.cell(i, 2, ee)
                    worksheet2.cell(i, 3, ee)
                else:
                    ee = sh.cell(row=i, column=j).value
                    worksheet2.cell(i, j, ee)

            if j == 2 or j == 3:
                # if type(c) == str and j == 2:
                d = len(cc.value)
                # print(d)
                if d < 4:
                    # print(c)
                    # print("Thevalueis", c)
                    ee = sh.cell(row=i, column=j).value
                    ff = sh.cell(row=i, column=4).value
                    # print(ee)
                    # print(ff)
                    worksheet2.cell(i, 4, ee)
                    worksheet2.cell(i, j, ff)
                else:

                    ee = sh.cell(row=i, column=j).value
                    worksheet2.cell(i, j, ee)
            else:
                ee = sh.cell(row=i, column=j).value
                # print(ee)
                worksheet2.cell(i, j, ee)

workbook2.save('result.xlsx')
full_data2 = pd.read_excel('result.xlsx', sheet_name='Sheet')
full_data2.iloc[:,:3] = full_data2.fillna(method='ffill')# 填充
full_data2.iloc[:,3] = full_data2.fillna('all')# 填充
print(full_data2)
full_data2.to_excel('done.xlsx', index=False)  # 不要索引

# print(ccc)